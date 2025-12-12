# streamlit_app_final3.py
import streamlit as st
import msal
import requests
import re
import os
import json
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, date, timedelta
from calendar import monthrange
from dateutil import parser as dateutil_parser
import pandas as pd

# -------------------------
# CONFIG (place secrets in .streamlit/secrets.toml)
# -------------------------
CLIENT_ID = st.secrets["client_id"]
CLIENT_SECRET = st.secrets["client_secret"]
TENANT_ID = st.secrets["tenant_id"]
REDIRECT_URI = st.secrets["redirect_uri"]

AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"
SCOPES = ["User.Read", "Mail.Read", "Files.ReadWrite"]
GRAPH = "https://graph.microsoft.com/v1.0"

# Optional: local template bytes (if you uploaded templates to runtime)
LOCAL_DSR_TEMPLATE = "/mnt/data/DSR_tracker.xlsx"   # optional - can be absent
LOCAL_WSR_TEMPLATE = "/mnt/data/WSR_tracker.xlsx"   # optional - can be absent

# Name of files stored in OneDrive root
ONEDRIVE_DSR_FILENAME = "DSR.xlsx"
ONEDRIVE_WSR_FILENAME = "WSR.xlsx"
PROCESSED_JSON_NAME = "processed_messages.json"

# Styling
HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # light blue
Y_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")      # light green
HEADER_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

# -------------------------
# MSAL helpers
# -------------------------
def build_app(cache=None):
    return msal.ConfidentialClientApplication(
        CLIENT_ID, authority=AUTHORITY, client_credential=CLIENT_SECRET, token_cache=cache
    )

def load_cache():
    cache = msal.SerializableTokenCache()
    if "token_cache" in st.session_state:
        cache.deserialize(st.session_state["token_cache"])
    return cache

def save_cache(cache):
    if cache.has_state_changed:
        st.session_state["token_cache"] = cache.serialize()

def get_token_silent():
    cache = load_cache()
    app = build_app(cache)
    accts = app.get_accounts()
    if accts:
        res = app.acquire_token_silent(SCOPES, account=accts[0])
        save_cache(cache)
        return res
    return None

# -------------------------
# Graph low-level helpers
# -------------------------
def graph_get_raw(url, token, params=None):
    headers = {"Authorization": f"Bearer {token}"}
    return requests.get(url, headers=headers, params=params)

def graph_get(url, token, params=None):
    r = graph_get_raw(url, token, params=params)
    if not r.ok:
        raise RuntimeError(f"Graph GET failed: {r.status_code} {r.text}")
    return r.json()

def graph_post(url, token, json_body=None):
    headers = {"Authorization": f"Bearer {token}"}
    r = requests.post(url, headers=headers, json=json_body)
    if not r.ok:
        raise RuntimeError(f"Graph POST failed: {r.status_code} {r.text}")
    return r.json()

def graph_put(url, token, data=None):
    headers = {"Authorization": f"Bearer {token}"}
    r = requests.put(url, headers=headers, data=data)
    if not r.ok:
        raise RuntimeError(f"Graph PUT failed: {r.status_code} {r.text}")
    return r.json()

# -------------------------
# OneDrive & Mail helpers
# -------------------------
def find_onedrive_file_id(token, filename):
    """
    Strict exact-name lookup in OneDrive root children (no fuzzy search).
    Returns item id or None.
    """
    url = f"{GRAPH}/me/drive/root/children"
    r = graph_get_raw(url, token)
    if r.status_code == 404:
        return None
    if not r.ok:
        raise RuntimeError(f"List children failed: {r.status_code} {r.text}")
    items = r.json().get("value", [])
    for it in items:
        if it.get("name", "").lower() == filename.lower():
            return it.get("id")
    return None

def create_onedrive_file_from_bytes(token, filename, file_bytes):
    url = f"{GRAPH}/me/drive/root:/{filename}:/content"
    headers = {"Authorization": f"Bearer {token}"}
    r = requests.put(url, headers=headers, data=file_bytes)
    if not r.ok:
        raise RuntimeError(f"Create file failed: {r.status_code} {r.text}")
    return r.json()

def download_onedrive_file_bytes(token, item_id):
    url = f"{GRAPH}/me/drive/items/{item_id}/content"
    headers = {"Authorization": f"Bearer {token}"}
    r = requests.get(url, headers=headers)
    if not r.ok:
        raise RuntimeError(f"Download file failed: {r.status_code} {r.text}")
    return r.content

def upload_onedrive_file_bytes(token, item_id, file_bytes):
    url = f"{GRAPH}/me/drive/items/{item_id}/content"
    headers = {"Authorization": f"Bearer {token}"}
    r = requests.put(url, headers=headers, data=file_bytes)
    if not r.ok:
        raise RuntimeError(f"Upload file failed: {r.status_code} {r.text}")
    return r.json()

def get_onedrive_item_metadata(token, item_id):
    url = f"{GRAPH}/me/drive/items/{item_id}"
    return graph_get(url, token)

# Mail
def get_inbox_messages(token, top=250):
    url = f"{GRAPH}/me/mailFolders/inbox/messages"
    params = {"$top": top, "$select": "id,subject,body,receivedDateTime,from,hasAttachments"}
    return graph_get(url, token, params).get("value", [])

# -------------------------
# Processed IDs (idempotency)
# -------------------------
def find_or_create_processed_json(token):
    fid = find_onedrive_file_id(token, PROCESSED_JSON_NAME)
    if fid:
        try:
            raw = download_onedrive_file_bytes(token, fid)
            obj = json.loads(raw.decode("utf-8"))
        except Exception:
            obj = {"processed_message_ids": []}
        return fid, obj
    initial = {"processed_message_ids": []}
    b = json.dumps(initial).encode("utf-8")
    meta = create_onedrive_file_from_bytes(token, PROCESSED_JSON_NAME, b)
    return meta.get("id"), initial

def save_processed_json(token, item_id, obj):
    b = json.dumps(obj, indent=2).encode("utf-8")
    upload_onedrive_file_bytes(token, item_id, b)

# -------------------------
# Date extraction helpers
# -------------------------
DATE_REGEXES = [
    r'(\d{1,2}(?:st|nd|rd|th)?\s+[A-Za-z]{3,9}\s+\d{4})',
    r'([A-Za-z]{3,9}\s+\d{1,2}(?:st|nd|rd|th)?,?\s+\d{4})',
    r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})',
    r'(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})',
    r'week\s+ending\s+([A-Za-z0-9\-\s,/]+)',
    r'(\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4}\s*[-–]\s*\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4})',
    r'from\s+(\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4})\s+to\s+(\d{1,2}\s+[A-Za-z]{3,9}\s+\d{4})'
]

def extract_date_strings(text):
    text = text or ""
    matches = []
    for rx in DATE_REGEXES:
        for m in re.finditer(rx, text, flags=re.IGNORECASE):
            matches.append(m.group(1))
            if m.lastindex and m.lastindex >= 2:
                matches.append(m.group(2))
    return matches

def parse_date_string(s):
    if not s:
        return None
    s = str(s).strip().strip("<>").replace(".", " ").replace(",", " ")
    try:
        d = dateutil_parser.parse(s, dayfirst=False, fuzzy=True).date()
        return d
    except Exception:
        pass
    fmts = ["%d %B %Y","%d %b %Y","%Y-%m-%d","%d-%m-%Y","%d/%m/%Y","%d/%m/%y","%B %d %Y","%b %d %Y"]
    for f in fmts:
        try:
            return datetime.strptime(s, f).date()
        except Exception:
            continue
    return None

# -------------------------
# DSR extraction from threads
# -------------------------
def extract_all_dsr_blocks(body_html):
    text = re.sub(r"<[^>]+>", "\n", body_html or "")
    splits = re.split(r"(?i)Daily Status Report", text)
    entries = []
    if len(splits) <= 1:
        return entries
    for s in splits[1:]:
        blk = "Daily Status Report " + s
        # date heuristics
        date_match = re.search(r"Daily Status Report.*?(\d{1,2}[^A-Za-z0-9\n]{0,2}\s*[A-Za-z0-9,\-\/\s]+?\d{4})", blk, flags=re.I)
        raw_date = None
        if date_match:
            raw_date = date_match.group(1)
        else:
            dm = re.search(r"for\s*[:\-]?\s*<?([\dA-Za-z\-/\s,]+?\d{2,4})>", blk, flags=re.I)
            if dm:
                raw_date = dm.group(1)
        parsed_date = parse_date_string(raw_date) if raw_date else None
        if not parsed_date:
            ds = extract_date_strings(blk)
            for sdate in ds:
                d = parse_date_string(sdate)
                if d:
                    parsed_date = d
                    break
        if not parsed_date:
            continue
        # project
        project = None
        mproj = re.search(r"Project\s*Name[:\s\-]+(.+)", blk, flags=re.I)
        if mproj:
            project = mproj.group(1).splitlines()[0].strip()
        else:
            lines = [ln.strip() for ln in blk.splitlines() if ln.strip()]
            if len(lines) >= 2:
                project = lines[1]
        # resource
        resource = None
        mres = re.search(r"Resource\s*Name[:\s\-]+(.+)", blk, flags=re.I)
        if mres:
            resource = mres.group(1).splitlines()[0].strip()
        if project and resource:
            entries.append({"project": project, "resource": resource, "date": parsed_date})
    return entries

# -------------------------
# WSR parsing
# -------------------------
def is_valid_wsr_subject(subject: str) -> bool:
    if not subject:
        return False

    s = subject.lower().strip()

    # Normalize dashes
    s = s.replace("–", "-").replace("—", "-")

    # Remove prefixes like "Re:", "Fw:", "Fwd:"
    s = re.sub(r'^(re|fw|fwd)\s*:\s*', '', s).strip()

    return (
        "wsr -" in s or
        "wsr-" in s or
        "weekly status report" in s
    )

def extract_wsr_project(subject: str):
    if not subject:
        return None

    s = subject.replace("–", "-").replace("—", "-").strip()
    s = re.sub(r'^(re|fw|fwd)\s*:\s*', '', s, flags=re.I).strip()

    # Expected formats:
    # WSR - Project
    # Weekly Status Report - Project
    parts = s.split("-", maxsplit=1)
    if len(parts) < 2:
        return None

    project = parts[1].strip()
    return project if project else None

def extract_wsr_date_range(body_html: str):
    if not body_html:
        return (None, None)

    # Clean HTML → plain text
    text = re.sub(r"<[^>]+>", " ", body_html)
    text = re.sub(r"\s+", " ", text)

    # Pattern 1: "from X to Y"
    m = re.search(r'from\s+(.+?)\s+to\s+(.+?\d{2,4})', text, flags=re.I)
    if m:
        s1 = parse_date_string(m.group(1))
        s2 = parse_date_string(m.group(2))
        if s1 and s2:
            return s1, s2

    # Pattern 2: "X - Y"
    m2 = re.search(r'(.+?\d{2,4})\s*[-–]\s*(.+?\d{2,4})', text)
    if m2:
        s1 = parse_date_string(m2.group(1))
        s2 = parse_date_string(m2.group(2))
        if s1 and s2:
            return s1, s2

    return (None, None)

# -------------------------
# Excel helpers
# -------------------------
def month_sheet_name(d: date):
    return d.strftime("%B %Y")

def ensure_month_sheet_and_days(wb, target_date: date):
    sheet_name = month_sheet_name(target_date)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        ws.cell(row=1, column=1, value="Project Name")
        ws.cell(row=1, column=2, value="Resource Name")
    year = target_date.year; month = target_date.month
    num_days = monthrange(year, month)[1]
    existing = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val:
            existing[str(val).strip()] = c
    for day in range(1, num_days+1):
        dt = date(year, month, day)
        label = dt.strftime("%d %B %Y")
        if label not in existing:
            col = ws.max_column + 1
            ws.cell(row=1, column=col, value=label)
            ws.cell(row=2, column=col, value=dt.strftime("%A"))
            existing[label] = col
    return ws, existing

def week_label_for_date(d: date):
    monday = d - timedelta(days=d.weekday())
    friday = monday + timedelta(days=4)
    return f"{monday.strftime('%d %b %Y')} - {friday.strftime('%d %b %Y')}"

def ensure_month_sheet_and_weeks(wb, any_date: date):
    sheet_name = month_sheet_name(any_date)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        ws.cell(row=1, column=1, value="Project Name")
    year = any_date.year; month = any_date.month
    first = date(year, month, 1)
    last = date(year, month, monthrange(year, month)[1])
    start_monday = first - timedelta(days=first.weekday())
    curr = start_monday
    existing = { (ws.cell(row=1, column=c).value or "").strip(): c for c in range(1, ws.max_column+1) }
    while curr <= last:
        mono = curr
        fri = mono + timedelta(days=4)
        label = f"{mono.strftime('%d %b %Y')} - {fri.strftime('%d %b %Y')}"
        if label not in existing:
            col = ws.max_column + 1
            ws.cell(row=1, column=col, value=label)
            existing[label] = col
        curr += timedelta(days=7)
    return ws, existing

def find_week_column_for_date(ws, target_date):
    for c in range(2, ws.max_column+1):
        header = ws.cell(row=1, column=c).value
        if not header: continue
        m = re.match(r'(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})\s*-\s*(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})', str(header))
        if m:
            start = parse_date_string(m.group(1))
            end = parse_date_string(m.group(2))
            if start and end and start <= target_date <= end:
                return c
    return None

def auto_fit_column_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            except Exception:
                pass
        adjusted_width = (max_len + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

def style_workbook_headers_and_ys(wb):
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # header row 1 style
        for c in range(1, ws.max_column+1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
        # style "Y" cells with light green
        for r in range(2, ws.max_row+1):
            for c in range(3, ws.max_column+1):
                cell = ws.cell(row=r, column=c)
                if cell.value == "Y":
                    cell.fill = Y_FILL
                    cell.alignment = CENTER_ALIGN

# -------------------------
# Workbook updates
# -------------------------
def update_dsr_wb_bytes(wb, dsr_records):
    processed = 0; duplicates = 0
    for rec in dsr_records:
        d = rec["date"]
        ws, headers = ensure_month_sheet_and_days(wb, d)
        project = rec["project"]; resource = rec["resource"]
        date_label = d.strftime("%d %B %Y")
        col_idx = headers.get(date_label)
        if not col_idx:
            col_idx = ws.max_column + 1
            ws.cell(row=1, column=col_idx, value=date_label)
        # find/create row
        row_idx = None
        for r in range(3, ws.max_row+1):
            if ws.cell(row=r, column=1).value == project and ws.cell(row=r, column=2).value == resource:
                row_idx = r; break
        if not row_idx:
            row_idx = ws.max_row + 1
            if row_idx < 3: row_idx = 3
            ws.cell(row=row_idx, column=1, value=project)
            ws.cell(row=row_idx, column=2, value=resource)
        if ws.cell(row=row_idx, column=col_idx).value == "Y":
            duplicates += 1
        else:
            ws.cell(row=row_idx, column=col_idx, value="Y")
            processed += 1
    style_workbook_headers_and_ys(wb)
    for s in wb.sheetnames:
        auto_fit_column_width(wb[s])
    bio = BytesIO(); wb.save(bio); return bio.getvalue(), processed, duplicates

def update_wsr_wb_bytes(wb, wsr_records):
    processed = 0; duplicates = 0
    for rec in wsr_records:
        d = rec["date"]
        ws, headers = ensure_month_sheet_and_weeks(wb, d)
        col_idx = find_week_column_for_date(ws, d)
        if not col_idx:
            label = week_label_for_date(d)
            if label not in headers:
                col_idx = ws.max_column + 1
                ws.cell(row=1, column=col_idx, value=label)
                headers[label] = col_idx
            else:
                col_idx = headers[label]
        project = rec["project"]
        row_idx = None
        for r in range(2, ws.max_row+1):
            if ws.cell(row=r, column=1).value == project:
                row_idx = r; break
        if not row_idx:
            row_idx = ws.max_row + 1
            if row_idx < 2: row_idx = 2
            ws.cell(row=row_idx, column=1, value=project)
        if ws.cell(row=row_idx, column=col_idx).value == "Y":
            duplicates += 1
        else:
            ws.cell(row=row_idx, column=col_idx, value="Y")
            processed += 1
    style_workbook_headers_and_ys(wb)
    for s in wb.sheetnames:
        auto_fit_column_width(wb[s])
    bio = BytesIO(); wb.save(bio); return bio.getvalue(), processed, duplicates

# -------------------------
# Minimal template creators (guaranteed)
# -------------------------
def make_minimal_dsr_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Project Name", "Resource Name"])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

def make_minimal_wsr_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Project Name"])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

def read_local_template_bytes(path):
    if os.path.exists(path):
        with open(path, "rb") as f:
            return f.read()
    return None

# -------------------------
# Ensure OneDrive file exists (robust)
# -------------------------
def ensure_onedrive_file(token, filename, local_template=None, minimal_fn=None):
    """
    Ensure exact filename exists (and is downloadable). Recreate if missing or broken.
    Returns (item_id, webUrl)
    """
    fid = find_onedrive_file_id(token, filename)
    if fid:
        try:
            _ = download_onedrive_file_bytes(token, fid)
            meta = get_onedrive_item_metadata(token, fid)
            return fid, meta.get("webUrl")
        except Exception:
            fid = None  # force recreate

    # prepare bytes (prefer local template if provided)
    b = None
    if local_template:
        b = read_local_template_bytes(local_template)
    if not b and minimal_fn:
        b = minimal_fn()
    if not b:
        raise RuntimeError("No template bytes available to create file.")
    meta = create_onedrive_file_from_bytes(token, filename, b)
    return meta.get("id"), meta.get("webUrl")

# -------------------------
# Preview parsing (exclude processed messages)
# -------------------------
def parse_inbox_messages_for_preview(messages, processed_ids):
    dsr_parsed = []
    wsr_parsed = []
    for m in messages:
        msg_id = m.get("id")
        if msg_id in processed_ids:
            continue
        subj = m.get("subject","") or ""
        body = (m.get("body") or {}).get("content","") or ""
        has_attach = m.get("hasAttachments", False)
        # DSR blocks
        dsr_blocks = extract_all_dsr_blocks(body)
        for blk in dsr_blocks:
            dsr_parsed.append({
                "message_id": msg_id,
                "from": (m.get("from") or {}).get("emailAddress", {}).get("address",""),
                "subject": subj,
                "project": blk.get("project"),
                "resource": blk.get("resource"),
                "date": blk.get("date")
            })
        # WSR
        # ---- Improved WSR detection ----
        # Attachment is compulsory
        if is_valid_wsr_subject(subj) and has_attach:

            project = extract_wsr_project(subj)
            start, end = extract_wsr_date_range(body)

            # MUST find project + start + end dates
            if project and start and end:
                wsr_parsed.append({
                    "message_id": msg_id,
                    "from": (m.get("from") or {}).get("emailAddress", {}).get("address", ""),
                    "subject": subj,
                    "project": project,
                    "start_date": start,
                    "end_date": end,
                    "date": start,     # representative week date
                    "hasAttachment": True
                })

    return dsr_parsed, wsr_parsed

# -------------------------
# UI
# -------------------------
st.set_page_config(page_title="DSR/WSR TRacker", layout="wide")
st.title("DSR / WSR Tracker")

# Auth handshake
params = st.query_params
if "code" in params:
    code = params["code"]
    cache = load_cache(); app = build_app(cache)
    try:
        result = app.acquire_token_by_authorization_code(code, SCOPES, redirect_uri=REDIRECT_URI)
        save_cache(cache)
    except Exception as e:
        st.error(f"Auth error: {e}")
    st.query_params.clear()

token_res = get_token_silent()
if not token_res:
    st.warning("Not signed in")
    if st.button("Sign in with Microsoft"):
        cache = load_cache(); app = build_app(cache)
        auth_url = app.get_authorization_request_url(SCOPES, redirect_uri=REDIRECT_URI)
        save_cache(cache)
        st.markdown(f"[Click here to sign in]({auth_url})")
    st.stop()

access_token = token_res["access_token"]
# st.success("Signed in as " + token_res.get("id_token_claims", {}).get("preferred_username",""))

# Fetch profile using /me
try:
    me = graph_get(f"{GRAPH}/me", access_token)
    user_email = me.get("mail") or me.get("userPrincipalName") or "Unknown user"
except Exception:
    user_email = "Unknown user"

st.success(f"Signed in as {user_email}")

col1, col2 = st.columns(2)

# ---- DSR ----
with col1:
    st.header("DSR")
    if st.button("Scan Inbox for DSR"):
        with st.spinner("Fetching inbox..."):
            try:
                proc_id, proc_obj = find_or_create_processed_json(access_token)
                processed_ids = set(proc_obj.get("processed_message_ids", []))
            except Exception as e:
                st.error(f"Failed to load processed ids: {e}")
                processed_ids = set()
                proc_id = None
                proc_obj = {"processed_message_ids": []}

            try:
                msgs = get_inbox_messages(access_token, top=300)
            except Exception as e:
                st.error(f"Failed fetching inbox: {e}")
                st.stop()

            dsr_preview, _ = parse_inbox_messages_for_preview(msgs, processed_ids)

            if not dsr_preview:
                st.info("No new DSR blocks parsed (all processed or none found).")
            else:
                df = pd.DataFrame(dsr_preview)
                if "date" in df.columns:
                    df["date"] = df["date"].apply(lambda x: x.isoformat() if isinstance(x, (date, datetime)) else x)
                st.session_state["df_dsr_preview"] = df
                st.success(f"Found {len(df)} new DSR entries")

    if st.session_state.get("df_dsr_preview") is not None:
        st.subheader("DSR Preview (editable)")
        df_edit = st.data_editor(st.session_state["df_dsr_preview"], num_rows="dynamic", width='stretch')
        st.session_state["df_dsr_preview"] = df_edit

        if st.button("Confirm & Update DSR.xlsx"):
            proc_id, proc_obj = find_or_create_processed_json(access_token)
            processed_ids = set(proc_obj.get("processed_message_ids", []))

            dsr_records = []
            for i, row in df_edit.iterrows():
                mid = row.get("message_id")
                if mid in processed_ids:
                    continue
                proj = row.get("project"); res = row.get("resource"); dt_raw = row.get("date")
                dt = parse_date_string(dt_raw) if isinstance(dt_raw, str) else None
                if not (proj and res and dt):
                    st.warning(f"Skipping row {i} - missing fields")
                    continue
                dsr_records.append({"project": proj, "resource": res, "date": dt, "message_id": mid})

            if not dsr_records:
                st.info("No valid DSR rows to process.")
            else:
                try:
                    fid, web = ensure_onedrive_file(access_token, ONEDRIVE_DSR_FILENAME, LOCAL_DSR_TEMPLATE, make_minimal_dsr_bytes)
                except Exception as e:
                    st.error(f"Failed ensuring {ONEDRIVE_DSR_FILENAME}: {e}"); fid = None; web = None

                if fid:
                    try:
                        b = download_onedrive_file_bytes(access_token, fid)
                        wb = load_workbook(BytesIO(b))
                        new_bytes, proc_count, dup_count = update_dsr_wb_bytes(wb, dsr_records)
                        upload_onedrive_file_bytes(access_token, fid, new_bytes)
                        st.success(f"DSR updated: processed={proc_count} duplicates={dup_count}")
                        if web:
                            st.markdown(f"**DSR file:** [{web}]({web})")
                        # mark processed ids
                        for r in dsr_records:
                            processed_ids.add(r["message_id"])
                        if proc_id is None:
                            meta = create_onedrive_file_from_bytes(access_token, PROCESSED_JSON_NAME, json.dumps({"processed_message_ids": list(processed_ids)}).encode("utf-8"))
                            proc_id = meta.get("id")
                        else:
                            save_processed_json(access_token, proc_id, {"processed_message_ids": list(processed_ids)})
                    except Exception as e:
                        st.error(f"Failed updating DSR.xlsx: {e}")

# ---- WSR ----
with col2:
    st.header("WSR")
    if st.button("Scan Inbox for WSR"):
        with st.spinner("Fetching inbox..."):
            try:
                proc_id, proc_obj = find_or_create_processed_json(access_token)
                processed_ids = set(proc_obj.get("processed_message_ids", []))
            except Exception as e:
                st.error(f"Failed to load processed ids: {e}")
                processed_ids = set()
                proc_id = None
                proc_obj = {"processed_message_ids": []}

            try:
                msgs = get_inbox_messages(access_token, top=300)
            except Exception as e:
                st.error(f"Failed fetching inbox: {e}")
                st.stop()

            _, wsr_preview = parse_inbox_messages_for_preview(msgs, processed_ids)

            if not wsr_preview:
                st.info("No new WSR candidate messages found.")
            else:
                dfw = pd.DataFrame(wsr_preview)
                if "date" in dfw.columns:
                    dfw["date"] = dfw["date"].apply(lambda x: x.isoformat() if isinstance(x, (date, datetime)) else x)
                if "start_date" in dfw.columns:
                    dfw["start_date"] = dfw["start_date"].apply(lambda x: x.isoformat() if isinstance(x, (date, datetime)) else x)
                if "end_date" in dfw.columns:
                    dfw["end_date"] = dfw["end_date"].apply(lambda x: x.isoformat() if isinstance(x, (date, datetime)) else x)
                st.session_state["df_wsr_preview"] = dfw
                st.success(f"Found {len(dfw)} new WSR entries")

    if st.session_state.get("df_wsr_preview") is not None:
        st.subheader("WSR Preview (editable)")
        dfw_edit = st.data_editor(st.session_state["df_wsr_preview"], num_rows="dynamic", width='stretch')
        st.session_state["df_wsr_preview"] = dfw_edit

        if st.button("Confirm & Update WSR.xlsx"):
            proc_id, proc_obj = find_or_create_processed_json(access_token)
            processed_ids = set(proc_obj.get("processed_message_ids", []))

            wsr_records = []
            for i, row in dfw_edit.iterrows():
                mid = row.get("message_id")
                if mid in processed_ids:
                    continue
                proj = row.get("project")
                sd_raw = row.get("start_date"); ed_raw = row.get("end_date")
                sd = parse_date_string(sd_raw) if isinstance(sd_raw, str) else None
                ed = parse_date_string(ed_raw) if isinstance(ed_raw, str) else None
                has_attach = row.get("hasAttachment", False)
                if not (proj and sd and ed and has_attach):
                    st.warning(f"Skipping row {i} - missing fields or attachment")
                    continue
                wsr_records.append({"project": proj, "date": sd, "message_id": mid})

            if not wsr_records:
                st.info("No valid WSR rows to process.")
            else:
                try:
                    fid, web = ensure_onedrive_file(access_token, ONEDRIVE_WSR_FILENAME, LOCAL_WSR_TEMPLATE, make_minimal_wsr_bytes)
                except Exception as e:
                    st.error(f"Failed ensuring {ONEDRIVE_WSR_FILENAME}: {e}"); fid = None; web = None

                if fid:
                    try:
                        b = download_onedrive_file_bytes(access_token, fid)
                        wb = load_workbook(BytesIO(b))
                        new_bytes, proc_count, dup_count = update_wsr_wb_bytes(wb, wsr_records)
                        upload_onedrive_file_bytes(access_token, fid, new_bytes)
                        st.success(f"WSR updated: processed={proc_count} duplicates={dup_count}")
                        if web:
                            st.markdown(f"**WSR file:** [{web}]({web})")
                        for r in wsr_records:
                            processed_ids.add(r["message_id"])
                        if proc_id is None:
                            meta = create_onedrive_file_from_bytes(access_token, PROCESSED_JSON_NAME, json.dumps({"processed_message_ids": list(processed_ids)}).encode("utf-8"))
                            proc_id = meta.get("id")
                        else:
                            save_processed_json(access_token, proc_id, {"processed_message_ids": list(processed_ids)})
                    except Exception as e:
                        st.error(f"Failed updating WSR.xlsx: {e}")

# End of app
